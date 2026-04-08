from __future__ import annotations

import calendar
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Point, ReceptionStat, WebUser
from app.web.deps import get_current_user, get_db, is_restricted_manager

TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

router = APIRouter(prefix="/reception", tags=["reception"])


@router.get("", response_class=HTMLResponse)
async def reception_index(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: WebUser = Depends(get_current_user),
    point_id: int = 0,
    year: int = 0,
    month: int = 0,
):
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    points_result = await db.execute(select(Point).where(Point.is_active == True).order_by(Point.name))
    points = points_result.scalars().all()

    # If no point selected, use first
    if not point_id and points:
        point_id = points[0].id

    # Days in month
    _, days_in_month = calendar.monthrange(year, month)
    dates = [date(year, month, d) for d in range(1, days_in_month + 1)]

    # Load stats for this point+month
    stats_result = await db.execute(
        select(ReceptionStat).where(
            ReceptionStat.point_id == point_id,
            ReceptionStat.stat_date >= date(year, month, 1),
            ReceptionStat.stat_date <= date(year, month, days_in_month),
        )
    )
    stats_by_date: dict[date, ReceptionStat] = {s.stat_date: s for s in stats_result.scalars().all()}

    # Build grid data
    grid = []
    for d in dates:
        s = stats_by_date.get(d)
        grid.append({
            "date": d,
            "day": d.day,
            "items_given": s.items_given if s else None,
            "clients_count": s.clients_count if s else None,
            "acceptance_amount": float(s.acceptance_amount) if s and s.acceptance_amount is not None else None,
        })

    # Month nav
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    month_names = ["Январь","Февраль","Март","Апрель","Май","Июнь",
                   "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    # 0=Пн … 6=Вс
    day_names = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]

    return templates.TemplateResponse(request, "reception/index.html", {
        "current_user": current_user,
        "active_page": "reception",
        "points": points,
        "point_id": point_id,
        "year": year,
        "month": month,
        "month_name": month_names[month - 1],
        "grid": grid,
        "prev_year": prev_year,
        "prev_month": prev_month,
        "next_year": next_year,
        "next_month": next_month,
        "today_str": date.today().strftime('%Y-%m-%d'),
        "day_names": day_names,
    })


@router.post("/save")
async def save_cell(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: WebUser = Depends(get_current_user),
):
    """AJAX endpoint: save a single cell value."""
    if is_restricted_manager(current_user):
        return JSONResponse({"error": "permission denied"}, status_code=403)

    data = await request.json()
    point_id = int(data.get("point_id", 0))
    field = str(data.get("field", ""))
    raw_value = str(data.get("value", "")).strip()
    try:
        stat_date = date.fromisoformat(str(data.get("date", "")))
    except (ValueError, TypeError):
        return JSONResponse({"error": "invalid date"}, status_code=400)

    if field not in ("items_given", "clients_count", "acceptance_amount"):
        return JSONResponse({"error": "invalid field"}, status_code=400)

    # Upsert
    stat = (await db.execute(
        select(ReceptionStat).where(
            ReceptionStat.point_id == point_id,
            ReceptionStat.stat_date == stat_date,
        )
    )).scalar_one_or_none()

    if not stat:
        stat = ReceptionStat(point_id=point_id, stat_date=stat_date)
        db.add(stat)

    if not raw_value:
        setattr(stat, field, None)
    elif field in ("items_given", "clients_count"):
        try:
            setattr(stat, field, int(raw_value))
        except ValueError:
            return JSONResponse({"error": "invalid integer"}, status_code=400)
    else:
        try:
            setattr(stat, field, Decimal(raw_value.replace(",", ".")))
        except InvalidOperation:
            return JSONResponse({"error": "invalid number"}, status_code=400)

    await db.commit()
    return JSONResponse({"ok": True})


# ---------------------------------------------------------------------------
# Chart data — daily values for a given period (month or week)
# ---------------------------------------------------------------------------

from datetime import timedelta

DAY_NAMES_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


@router.get("/chart-data")
async def reception_chart_data(
    db: AsyncSession = Depends(get_db),
    current_user: WebUser = Depends(get_current_user),
    point_id: int = 0,
    period: str = "month",   # "month" | "week"
    year: int = 0,
    month: int = 0,
):
    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    if period == "week":
        # Last 7 days ending today
        d_end   = today
        d_start = today - timedelta(days=6)
    else:
        # Full calendar month
        _, days_in_month = calendar.monthrange(year, month)
        d_start = date(year, month, 1)
        d_end   = date(year, month, days_in_month)

    stats = (await db.execute(
        select(ReceptionStat).where(
            ReceptionStat.point_id == point_id,
            ReceptionStat.stat_date >= d_start,
            ReceptionStat.stat_date <= d_end,
        )
    )).scalars().all()
    stats_by_date = {s.stat_date: s for s in stats}

    results = []
    cur = d_start
    while cur <= d_end:
        s = stats_by_date.get(cur)
        if period == "week":
            label = f"{cur.day:02d}.{cur.month:02d} {DAY_NAMES_RU[cur.weekday()]}"
        else:
            label = f"{cur.day} {DAY_NAMES_RU[cur.weekday()]}"
        results.append({
            "label":          label,
            "items_given":    s.items_given if s and s.items_given is not None else None,
            "clients_count":  s.clients_count if s and s.clients_count is not None else None,
            "acceptance":     float(s.acceptance_amount) if s and s.acceptance_amount is not None else None,
        })
        cur += timedelta(days=1)

    return JSONResponse(results)


# ---------------------------------------------------------------------------
# Excel export — current month table
# ---------------------------------------------------------------------------

@router.get("/export")
async def export_excel(
    db: AsyncSession = Depends(get_db),
    current_user: WebUser = Depends(get_current_user),
    point_id: int = 0,
    year: int = 0,
    month: int = 0,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    if not year:
        year = today.year
    if not month:
        month = today.month

    point = (await db.execute(select(Point).where(Point.id == point_id))).scalar_one_or_none()
    point_name = point.name if point else f"point_{point_id}"

    _, days_in_month = calendar.monthrange(year, month)
    dates = [date(year, month, d) for d in range(1, days_in_month + 1)]

    stats_result = await db.execute(
        select(ReceptionStat).where(
            ReceptionStat.point_id == point_id,
            ReceptionStat.stat_date >= date(year, month, 1),
            ReceptionStat.stat_date <= date(year, month, days_in_month),
        )
    )
    stats_by_date = {s.stat_date: s for s in stats_result.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    month_names_full = ["Январь","Февраль","Март","Апрель","Май","Июнь",
                        "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    ws.title = f"{month_names_full[month - 1]} {year}"

    # Styles
    hdr_fill  = PatternFill("solid", fgColor="1E293B")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    lbl_fill  = PatternFill("solid", fgColor="F8FAFC")
    lbl_font  = Font(bold=True, size=10)
    sum_fill  = PatternFill("solid", fgColor="EFF6FF")
    avg_fill  = PatternFill("solid", fgColor="F0FDF4")
    warn_fill = PatternFill("solid", fgColor="FEF9C3")
    alrt_fill = PatternFill("solid", fgColor="FEE2E2")
    neg_fill  = PatternFill("solid", fgColor="FEE2E2")
    neg_font  = Font(color="B91C1C", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    center = Alignment(horizontal='center', vertical='center')

    day_names_ru = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 3)
    title_cell = ws.cell(1, 1, f"Статистика приёмки — {point_name} — {month_names_full[month-1]} {year}")
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = center

    # Header row (row 2): label | 1…N | Итого | Среднее/день
    ws.cell(2, 1, "Показатель").fill = hdr_fill
    ws.cell(2, 1).font = hdr_font
    ws.cell(2, 1).alignment = center
    ws.column_dimensions['A'].width = 22

    for idx, d in enumerate(dates, start=2):
        col = idx
        cell = ws.cell(2, col, f"{d.day}\n{day_names_ru[d.weekday()]}")
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 6

    sum_col = days_in_month + 2
    avg_col = days_in_month + 3
    ws.cell(2, sum_col, "Итого").fill = hdr_fill
    ws.cell(2, sum_col).font = hdr_font
    ws.cell(2, sum_col).alignment = center
    ws.column_dimensions[get_column_letter(sum_col)].width = 9
    ws.cell(2, avg_col, "Среднее\n/день").fill = hdr_fill
    ws.cell(2, avg_col).font = hdr_font
    ws.cell(2, avg_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(avg_col)].width = 10

    ws.row_dimensions[2].height = 30

    # Data rows
    row_defs = [
        ("items_given",        "Товаров отдали",  3),
        ("clients_count",      "Клиентов в день", 4),
        ("acceptance_amount",  "Приёмка, ₽",      5),
    ]
    for field, label, row in row_defs:
        ws.cell(row, 1, label).fill = lbl_fill
        ws.cell(row, 1).font = lbl_font
        ws.cell(row, 1).border = thin_border

        vals = []
        for idx, d in enumerate(dates, start=2):
            s = stats_by_date.get(d)
            raw = None
            if s:
                if field == "acceptance_amount":
                    raw = float(s.acceptance_amount) if s.acceptance_amount is not None else None
                else:
                    raw = getattr(s, field)
            cell = ws.cell(row, idx)
            if raw is not None:
                cell.value = raw
                vals.append(raw)
                # Coloring
                if field == "items_given":
                    if raw > 500:
                        cell.fill = alrt_fill
                    elif raw > 350:
                        cell.fill = warn_fill
                elif field == "acceptance_amount" and raw < 0:
                    cell.fill = neg_fill
                    cell.font = neg_font
            cell.alignment = center
            cell.border = thin_border
            if field == "acceptance_amount":
                cell.number_format = '#,##0.00'

        # Итого
        total = sum(vals) if vals else None
        sum_cell = ws.cell(row, sum_col)
        sum_cell.value = total
        sum_cell.fill = sum_fill
        sum_cell.font = Font(bold=True, size=10)
        sum_cell.alignment = center
        sum_cell.border = thin_border
        if field == "acceptance_amount":
            sum_cell.number_format = '#,##0.00'

        # Среднее/день
        avg_cell = ws.cell(row, avg_col)
        if vals and field != "acceptance_amount":
            avg_cell.value = round(sum(vals) / len(vals), 1)
            avg_cell.fill = avg_fill
            avg_cell.font = Font(bold=True, size=10)
        else:
            avg_cell.value = None
        avg_cell.alignment = center
        avg_cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"reception_{point_id}_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
