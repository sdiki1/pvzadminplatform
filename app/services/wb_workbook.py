from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.config import Settings
from app.services.google_sheets import MainStatsRow
from app.utils.parsing import parse_date, parse_decimal


class WBWorkbookService:
    def __init__(self, settings: Settings):
        self.settings = settings

    @property
    def enabled(self) -> bool:
        return bool(self.settings.wb_workbook_file)

    def fetch_main_stats(self) -> list[MainStatsRow]:
        if not self.enabled:
            return []

        path = Path(self.settings.wb_workbook_file)
        if not path.exists():
            return []

        wb = load_workbook(path, data_only=True, read_only=True)
        sheet_name = self.settings.wb_workbook_stats_sheet
        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        out: list[MainStatsRow] = []

        current_point: str | None = None
        date_row: int | None = None
        issued_row: int | None = None
        acceptance_row: int | None = None

        for row_idx in range(1, ws.max_row + 1):
            raw = ws.cell(row=row_idx, column=1).value
            label = str(raw).strip().lower() if raw is not None else ""

            if label.startswith("пвз"):
                current_point = str(raw).strip()
                date_row = None
                issued_row = None
                acceptance_row = None
                continue

            if "дата" == label:
                date_row = row_idx
                continue

            if "товаров отдали" in label:
                issued_row = row_idx
                continue

            if "статистика приемки" in label or "статистика приёмки" in label:
                acceptance_row = row_idx

                if not (current_point and date_row and issued_row and acceptance_row):
                    continue

                for col_idx in range(2, ws.max_column + 1):
                    d = parse_date(ws.cell(row=date_row, column=col_idx).value)
                    if not d:
                        continue

                    issued = int(parse_decimal(ws.cell(row=issued_row, column=col_idx).value))
                    acceptance = parse_decimal(ws.cell(row=acceptance_row, column=col_idx).value)

                    if issued == 0 and acceptance == 0:
                        # Пустые/нерабочие даты не импортируем.
                        continue

                    payload = {
                        "source": str(path),
                        "sheet": sheet_name,
                        "row": row_idx,
                        "col": col_idx,
                        "point": current_point,
                    }
                    out.append(
                        MainStatsRow(
                            record_date=d,
                            point_name=current_point,
                            manager_name=None,
                            acceptance_amount_rub=acceptance,
                            issued_items_count=max(0, issued),
                            tickets_count=0,
                            raw_payload=json.dumps(payload, ensure_ascii=False),
                        )
                    )

        return out
