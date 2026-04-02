from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import unicodedata

from openpyxl import Workbook

from app.db.models import Expense, PayrollItem, Point, Shift, User
from app.services.payroll import EmployeePayrollBreakdown


class ReportService:
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_payroll_summary_xlsx(
        self,
        period_start: date,
        period_end: date,
        rows: list[EmployeePayrollBreakdown],
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll"

        ws.append(["Период", f"{period_start:%d.%m.%Y} - {period_end:%d.%m.%Y}"])
        ws.append([])
        ws.append(
            [
                "Сотрудник",
                "Смены",
                "Часы",
                "База",
                "Скорость приемки",
                "Рейтинг",
                "Бонус выдача",
                "Резерв",
                "Резервный выход",
                "Зависшие",
                "Подмена товара",
                "Брак товара",
                "Удержания всего",
                "Бонус менеджера",
                "Корректировки",
                "Подытог",
                "Долг/Переплата ДС",
                "Итого",
            ]
        )

        for row in rows:
            ws.append(
                [
                    row.user.full_name,
                    row.shifts_count,
                    float(row.hours_total),
                    float(row.base_amount_rub),
                    float(row.motivation_amount_rub),
                    float(row.rating_bonus_rub),
                    float(row.issued_bonus_rub),
                    float(row.reserve_bonus_rub),
                    float(row.substitution_bonus_rub),
                    float(row.stuck_deduction_rub),
                    float(row.substitution_deduction_rub),
                    float(row.defect_deduction_rub),
                    float(row.dispute_deduction_rub),
                    float(row.manager_bonus_rub),
                    float(row.adjustments_rub),
                    float(row.subtotal_amount_rub),
                    float(row.debt_adjustment_rub),
                    float(row.total_amount_rub),
                ]
            )

        path = self.export_dir / f"payroll_summary_{period_start}_{period_end}.xlsx"
        wb.save(path)
        return path

    def export_employee_payroll_sheets(
        self,
        period_start: date,
        period_end: date,
        rows: list[EmployeePayrollBreakdown],
    ) -> Path:
        wb = Workbook()
        first = True

        for row in rows:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = (row.user.full_name[:28] or "Employee").replace("/", " ")

            ws.append([row.user.full_name])
            ws.append(["Период начислений", "", "", f"{period_start:%d.%m.%Y} - {period_end:%d.%m.%Y}"])
            ws.append([])
            ws.append(["Оклад", "", row.shifts_count, float(row.base_amount_rub)])
            ws.append(["Премия за скорость приёмки", "", "", float(row.motivation_amount_rub)])
            ws.append(["Премия за рейтинг, оценки клиентов", "", "", float(row.rating_bonus_rub)])
            ws.append(["Премия за выдачу", "", "", float(row.issued_bonus_rub)])
            ws.append(["Резервные дежурства", "", "", float(row.reserve_bonus_rub)])
            ws.append(["Резервные выходы", "", "", float(row.substitution_bonus_rub)])
            ws.append(["Зависшие товары", "", "", float(row.stuck_deduction_rub)])
            ws.append(["Подмена товара", "", "", float(row.substitution_deduction_rub)])
            ws.append(["Брак товара", "", "", float(row.defect_deduction_rub)])
            ws.append(["Удержания по товарам (не оспорено)", "", "", float(row.dispute_deduction_rub)])
            ws.append(["Доп. выплаты менеджера", "", "", float(row.manager_bonus_rub)])
            ws.append(["Премия / удержание руководства", "", "", float(row.adjustments_rub)])
            ws.append(["Подытог (без ДС)", "", "", float(row.subtotal_amount_rub)])
            ws.append(["Долг / Переплата ДС", "", "", float(row.debt_adjustment_rub)])
            ws.append(["Итого", "", "", float(row.total_amount_rub)])

        path = self.export_dir / f"payroll_sheets_{period_start}_{period_end}.xlsx"
        wb.save(path)
        return path

    @staticmethod
    def _num(value: object) -> float:
        if value is None:
            return 0.0
        if isinstance(value, Decimal):
            return float(value)
        try:
            return float(value)
        except Exception:
            return 0.0

    @staticmethod
    def _safe_filename(value: str) -> str:
        ascii_value = (
            unicodedata.normalize("NFKD", value.strip())
            .encode("ascii", "ignore")
            .decode("ascii")
        )
        cleaned = "".join(ch if ch.isalnum() or ch in ("_", "-", ".") else "_" for ch in ascii_value)
        cleaned = cleaned.strip("._")
        return cleaned or "employee"

    def export_employee_sheet_xlsx(
        self,
        *,
        run_id: int,
        item_id: int,
        employee_name: str,
        period_start: date,
        period_end: date,
        payout_day: int,
        item: PayrollItem,
        view_mode: str,
        details: dict,
        manager_bonus_3_per_ticket: int,
        reserve_duty_bonus_rub: int,
        substitution_bonus_rub: int,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"

        ws.append(["Расчётный лист"])
        ws.append(["Сотрудник", employee_name])
        ws.append(["Период", f"{period_start:%d.%m.%Y} - {period_end:%d.%m.%Y}"])
        ws.append(["День выплаты", payout_day])
        ws.append(["Режим", "Полная версия" if view_mode == "full" else "Краткая сводка"])
        ws.append([])

        summary_rows = [
            ("Количество смен", item.shifts_count),
            ("Отработано часов", self._num(item.hours_total)),
            ("Оклад (базовая часть)", self._num(item.base_amount_rub)),
            ("Премия за скорость приёмки", self._num(item.motivation_amount_rub)),
            ("Премия за рейтинг, оценки клиентов", self._num(item.rating_bonus_rub)),
            ("Премия за выдачу", self._num(item.issued_bonus_rub)),
            ("Резервные дежурства", self._num(item.reserve_bonus_rub)),
            ("Резервный выход", self._num(item.substitution_bonus_rub)),
            ("Зависшие товары", -self._num(item.stuck_deduction_rub)),
            ("Подмена товара", -self._num(item.substitution_deduction_rub)),
            ("Брак товара", -self._num(item.defect_deduction_rub)),
            ("Удержания по товарам (не оспорено)", -self._num(item.dispute_deduction_rub)),
            ("Доп. выплаты менеджера", self._num(item.manager_bonus_rub)),
            ("Премия / удержание руководства", self._num(item.adjustments_rub)),
            ("Подытог (без ДС)", self._num(item.total_amount_rub) - self._num(item.debt_adjustment_rub)),
            ("Долг / Переплата ДС", self._num(item.debt_adjustment_rub)),
            ("ИТОГО К ВЫПЛАТЕ", self._num(item.total_amount_rub)),
        ]
        for label, value in summary_rows:
            ws.append([label, value])

        if view_mode == "full":
            ws.append([])
            ws.append(["Полная детализация"])
            ws.append([])

            ws.append(["1. Смены и стоимость"])
            ws.append(["Дата", "ПВЗ", "Часы", "Тип", "Формула", "Стоимость", "Подмена"])
            for row in details.get("shift_rows", []):
                ws.append(
                    [
                        row["shift_date"].strftime("%d.%m.%Y"),
                        row["point_name"],
                        self._num(row.get("hours")),
                        row.get("basis", ""),
                        row.get("formula", ""),
                        self._num(row.get("amount_rub")),
                        "Да" if row.get("is_substitution") else "Нет",
                    ]
                )
            ws.append(["", "", "", "", "Итого по сменам", self._num(details.get("shift_amount_total")), ""])
            ws.append([])

            ws.append([f"2. Резервные дежурства (+{reserve_duty_bonus_rub} ₽)"])
            ws.append(["Дата", "ПВЗ", "Начисление"])
            for row in details.get("reserve_rows", []):
                ws.append([row["shift_date"].strftime("%d.%m.%Y"), row["point_name"], self._num(row["amount_rub"])])
            ws.append(["", "Итого резерв", self._num(details.get("reserve_amount_total"))])
            ws.append([])

            ws.append([f"3. Резервные выходы (+{substitution_bonus_rub} ₽)"])
            ws.append(["Дата", "ПВЗ", "Начисление"])
            for row in details.get("substitution_rows", []):
                ws.append([row["shift_date"].strftime("%d.%m.%Y"), row["point_name"], self._num(row["amount_rub"])])
            ws.append(["", "Итого подмена", self._num(details.get("substitution_amount_total"))])
            ws.append([])

            ws.append(["4. Списания из оспариваний"])
            ws.append(["Дата", "ПВЗ", "Тип", "Сумма", "ШК", "Тикет", "Статус", "Описание", "Ссылка"])
            for row in details.get("appeal_rows", []):
                ws.append(
                    [
                        row["case_date"].strftime("%d.%m.%Y"),
                        row["point_name"],
                        row["type_label"],
                        -self._num(row["amount_rub"]),
                        row["barcode"] or "",
                        row["ticket_number"] or "",
                        row["status_label"],
                        row["description"] or "",
                        f"/appeals/{row['id']}",
                    ]
                )
            totals = details.get("appeal_totals", {})
            ws.append(["", "", "Зависшие", -self._num(totals.get("stuck")), "", "", "", "", ""])
            ws.append(["", "", "Подмена товара", -self._num(totals.get("substitution")), "", "", "", "", ""])
            ws.append(["", "", "Брак товара", -self._num(totals.get("defect")), "", "", "", "", ""])
            ws.append(["", "", "Прочие списания", -self._num(totals.get("other")), "", "", "", "", ""])
            ws.append([])

            details_map = details.get("details", {})
            ws.append(["5. Премия за выдачу (детали)"])
            ws.append(["Показатель", "Значение"])
            ws.append(["Количество выданных товаров", details_map.get("issued_items", "0")])
            ws.append(["Автоматическая премия за выдачу", details_map.get("issued_bonus_auto_rub", "0.00")])
            ws.append(["Итоговая премия за выдачу", self._num(item.issued_bonus_rub)])
            ws.append([])

            ws.append([f"6. Премия/удержание руководства (бонус type3: {manager_bonus_3_per_ticket} ₽/тикет)"])
            ws.append(["Тип", "Комментарий", "Сумма"])
            for row in details.get("adjustment_rows", []):
                ws.append([row["adjustment_type"], row["comment"] or "", self._num(row["amount_rub"])])

        filename = (
            f"payroll_sheet_run{run_id}_item{item_id}_{self._safe_filename(employee_name)}_{view_mode}.xlsx"
        )
        path = self.export_dir / filename
        wb.save(path)
        return path

    def export_employee_sheet_pdf(
        self,
        *,
        run_id: int,
        item_id: int,
        employee_name: str,
        period_start: date,
        period_end: date,
        payout_day: int,
        item: PayrollItem,
        view_mode: str,
        details: dict,
        manager_bonus_3_per_ticket: int,
        reserve_duty_bonus_rub: int,
        substitution_bonus_rub: int,
    ) -> Path:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except Exception as exc:
            raise RuntimeError("PDF export requires 'reportlab' package in environment") from exc

        font_name = None
        font_candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Times New Roman.ttf",
        ]
        for candidate in font_candidates:
            if Path(candidate).exists():
                try:
                    pdfmetrics.getFont("PVZSans")
                except Exception:
                    pdfmetrics.registerFont(TTFont("PVZSans", candidate))
                font_name = "PVZSans"
                break

        if not font_name:
            raise RuntimeError(
                "No Unicode font found for PDF export. "
                "Install fonts-dejavu-core in container or provide DejaVuSans.ttf."
            )

        page_size = landscape(A4) if view_mode == "full" else A4
        filename = f"payroll_sheet_run{run_id}_item{item_id}_{self._safe_filename(employee_name)}_{view_mode}.pdf"
        path = self.export_dir / filename
        doc = SimpleDocTemplate(
            str(path),
            pagesize=page_size,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "pvzTitle",
            parent=styles["Heading3"],
            fontName=font_name,
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#0f172a"),
        )
        normal_style = ParagraphStyle(
            "pvzNormal",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#334155"),
        )
        section_style = ParagraphStyle(
            "pvzSection",
            parent=styles["Heading4"],
            fontName=font_name,
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#1d4ed8"),
        )

        def _table(
            rows: list[list[object]],
            repeat: int = 1,
            highlight_last_row: bool = False,
            detect_negative: bool = True,
        ) -> Table:
            table = Table(rows, repeatRows=repeat)
            style_cmds: list[tuple] = [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
                ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
            if len(rows) > 1:
                for row_idx in range(1, len(rows)):
                    if row_idx % 2 == 0:
                        style_cmds.append(
                            ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#f8fafc"))
                        )
            if highlight_last_row and len(rows) > 1:
                style_cmds.extend(
                    [
                        ("BACKGROUND", (0, len(rows) - 1), (-1, len(rows) - 1), colors.HexColor("#dcfce7")),
                        ("TEXTCOLOR", (0, len(rows) - 1), (-1, len(rows) - 1), colors.HexColor("#14532d")),
                        ("FONTNAME", (0, len(rows) - 1), (-1, len(rows) - 1), font_name),
                    ]
                )
            if detect_negative:
                for row_idx in range(1, len(rows)):
                    for col_idx, cell_value in enumerate(rows[row_idx]):
                        cell_text = str(cell_value).strip().replace(" ", "")
                        if cell_text.startswith("-"):
                            style_cmds.append(
                                ("TEXTCOLOR", (col_idx, row_idx), (col_idx, row_idx), colors.HexColor("#b91c1c"))
                            )
            table.setStyle(TableStyle(style_cmds))
            return table

        story = []
        story.append(Paragraph("Расчётный лист", title_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"Сотрудник: {employee_name}", normal_style))
        story.append(
            Paragraph(
                f"Период: {period_start:%d.%m.%Y} - {period_end:%d.%m.%Y}; Выплата: {payout_day}-го числа",
                normal_style,
            )
        )
        story.append(Paragraph(f"Режим: {'Полная версия' if view_mode == 'full' else 'Краткая сводка'}", normal_style))
        story.append(Spacer(1, 10))

        summary_rows = [
            ["Показатель", "Значение"],
            ["Количество смен", item.shifts_count],
            ["Отработано часов", self._num(item.hours_total)],
            ["Оклад (базовая часть)", self._num(item.base_amount_rub)],
            ["Премия за скорость приёмки", self._num(item.motivation_amount_rub)],
            ["Премия за рейтинг", self._num(item.rating_bonus_rub)],
            ["Премия за выдачу", self._num(item.issued_bonus_rub)],
            ["Резервные дежурства", self._num(item.reserve_bonus_rub)],
            ["Резервный выход", self._num(item.substitution_bonus_rub)],
            ["Зависшие товары", -self._num(item.stuck_deduction_rub)],
            ["Подмена товара", -self._num(item.substitution_deduction_rub)],
            ["Брак товара", -self._num(item.defect_deduction_rub)],
            ["Удержания по товарам", -self._num(item.dispute_deduction_rub)],
            ["Доп. выплаты менеджера", self._num(item.manager_bonus_rub)],
            ["Премия / удержание руководства", self._num(item.adjustments_rub)],
            ["Подытог (без ДС)", self._num(item.total_amount_rub) - self._num(item.debt_adjustment_rub)],
            ["Долг / Переплата ДС", self._num(item.debt_adjustment_rub)],
            ["ИТОГО К ВЫПЛАТЕ", self._num(item.total_amount_rub)],
        ]
        story.append(_table(summary_rows, highlight_last_row=True))

        if view_mode == "full":
            story.append(Spacer(1, 10))
            story.append(Paragraph("Полная детализация", section_style))
            story.append(Spacer(1, 6))

            shift_rows = [["Дата", "ПВЗ", "Часы", "Тип", "Формула", "Стоимость", "Подмена"]]
            for row in details.get("shift_rows", []):
                shift_rows.append(
                    [
                        row["shift_date"].strftime("%d.%m.%Y"),
                        row["point_name"],
                        f"{self._num(row.get('hours')):.2f}",
                        row.get("basis", ""),
                        row.get("formula", ""),
                        f"{self._num(row.get('amount_rub')):.2f}",
                        "Да" if row.get("is_substitution") else "Нет",
                    ]
                )
            if len(shift_rows) == 1:
                shift_rows.append(["—", "Нет данных", "", "", "", "", ""])
            story.append(Paragraph("1. Смены и стоимость", section_style))
            story.append(_table(shift_rows))
            story.append(Spacer(1, 6))

            appeal_rows = [["Дата", "ПВЗ", "Тип", "Сумма", "ШК", "Тикет", "Статус", "Описание", "Ссылка"]]
            for row in details.get("appeal_rows", []):
                appeal_rows.append(
                    [
                        row["case_date"].strftime("%d.%m.%Y"),
                        row["point_name"],
                        row["type_label"],
                        f"-{self._num(row['amount_rub']):.2f}",
                        row["barcode"] or "—",
                        row["ticket_number"] or "—",
                        row["status_label"],
                        row["description"] or "—",
                        f"/appeals/{row['id']}",
                    ]
                )
            if len(appeal_rows) == 1:
                appeal_rows.append(["—", "Нет списаний", "", "", "", "", "", "", ""])
            story.append(Paragraph("2. Списания из оспариваний", section_style))
            story.append(_table(appeal_rows))
            story.append(Spacer(1, 6))

            details_map = details.get("details", {})
            issue_rows = [
                ["Показатель", "Значение"],
                ["Количество выданных товаров", str(details_map.get("issued_items", "0"))],
                ["Автоматическая премия за выдачу", str(details_map.get("issued_bonus_auto_rub", "0.00"))],
                ["Итоговая премия за выдачу", f"{self._num(item.issued_bonus_rub):.2f}"],
            ]
            story.append(Paragraph("3. Премия за выдачу (детали)", section_style))
            story.append(_table(issue_rows, detect_negative=False))
            story.append(Spacer(1, 6))

            adjustment_rows = [["Тип", "Комментарий", "Сумма"]]
            for row in details.get("adjustment_rows", []):
                adjustment_rows.append(
                    [
                        row["adjustment_type"],
                        row["comment"] or "—",
                        f"{self._num(row['amount_rub']):.2f}",
                    ]
                )
            if len(adjustment_rows) == 1:
                adjustment_rows.append(["—", "Нет ручных корректировок", "0.00"])
            story.append(
                Paragraph(
                    (
                        "4. Премия/удержание руководства "
                        f"(бонус type3: {manager_bonus_3_per_ticket} ₽/тикет; "
                        f"резерв: +{reserve_duty_bonus_rub}; подмена: +{substitution_bonus_rub})"
                    ),
                    section_style,
                )
            )
            story.append(_table(adjustment_rows))

        doc.build(story)
        return path

    def export_shifts_csv(
        self,
        period_start: date,
        period_end: date,
        shifts: list[Shift],
        users_by_id: dict[int, User],
        points_by_id: dict[int, Point],
    ) -> Path:
        path = self.export_dir / f"shifts_{period_start}_{period_end}.csv"
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(
                [
                    "id",
                    "date",
                    "employee",
                    "point",
                    "opened_at",
                    "closed_at",
                    "duration_min",
                    "open_distance_m",
                    "close_distance_m",
                ]
            )
            for s in shifts:
                writer.writerow(
                    [
                        s.id,
                        s.shift_date.isoformat(),
                        users_by_id.get(s.user_id).full_name if s.user_id in users_by_id else s.user_id,
                        points_by_id.get(s.point_id).name if s.point_id in points_by_id else s.point_id,
                        s.opened_at.isoformat() if s.opened_at else "",
                        s.closed_at.isoformat() if s.closed_at else "",
                        s.duration_minutes or 0,
                        float(s.open_distance_m or 0),
                        float(s.close_distance_m or 0),
                    ]
                )
        return path

    def export_expenses_csv(self, period_start: date, period_end: date, expenses: list[Expense], points_by_id: dict[int, Point]) -> Path:
        path = self.export_dir / f"expenses_{period_start}_{period_end}.csv"
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["date", "point", "category", "amount_rub", "description"])
            for e in expenses:
                writer.writerow(
                    [
                        e.expense_date.isoformat(),
                        points_by_id.get(e.point_id).name if e.point_id in points_by_id else e.point_id,
                        e.category,
                        float(e.amount_rub),
                        e.description or "",
                    ]
                )
        return path
