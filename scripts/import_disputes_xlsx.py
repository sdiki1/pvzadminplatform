from __future__ import annotations

import argparse
import asyncio
import json
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.db.models import MotivationRecord, MotivationSource, Point, User
from app.db.session import SessionLocal, init_db
from app.utils.parsing import normalize_text, parse_date, parse_decimal

NICKNAMES = {
    "дима": "дмитрий",
    "даня": "даниил",
    "ксюша": "ксения",
    "даша": "дария",
    "дарья": "дария",
    "вика": "виктория",
    "варя": "варвара",
    "катя": "екатерина",
    "настя": "анастасия",
}


@dataclass
class DisputeXlsxRow:
    record_date: date
    point_key: str | None
    manager_name: str | None
    status: str | None
    amount_rub: Decimal
    raw_payload: str


def _norm_key(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^\w\sа-яА-ЯёЁ]", " ", text, flags=re.UNICODE)
    text = normalize_text(text.replace("улица", "ул").replace("ул.", "ул"))
    return text


def _extract_address_key(value: str) -> str:
    text = _norm_key(value)
    parts = [p for p in text.split() if p not in {"пвз", "ул", "улица", "г", "город", "wb", "ozon", "озон"}]
    return " ".join(parts)


def _person_tokens(name: str | None) -> list[str]:
    if not name:
        return []
    t = normalize_text(name).replace("ё", "е")
    t = re.sub(r"[^а-яa-z\s\-]", " ", t, flags=re.UNICODE)
    tokens = [x for x in t.split() if x]
    out: list[str] = []
    for tok in tokens:
        out.append(NICKNAMES.get(tok, tok))
    return out


def _person_keys(name: str | None) -> set[str]:
    tokens = _person_tokens(name)
    if not tokens:
        return set()
    keys = {" ".join(tokens), tokens[0], tokens[-1]}
    if len(tokens) >= 2:
        keys.add(f"{tokens[0]} {tokens[1]}")
        keys.add(f"{tokens[1]} {tokens[0]}")
        keys.add(f"{tokens[0]} {tokens[-1]}")
        keys.add(f"{tokens[-1]} {tokens[0]}")
    return keys


def _build_header_map(header_row: list[object | None]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        key = normalize_text(str(value))
        if key:
            out[key] = idx
    return out


def _header_index(header_map: dict[str, int], aliases: list[str], default_idx: int | None = None) -> int | None:
    for alias in aliases:
        idx = header_map.get(normalize_text(alias))
        if idx is not None:
            return idx
    return default_idx


def _cell_text(value: object | None) -> str | None:
    if value is None:
        return None
    txt = str(value).strip()
    return txt or None


def _parse_xlsx(path: Path, sheet_name: str | None) -> list[DisputeXlsxRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = list(next(rows_iter, ()))
    header_map = _build_header_map(header)

    date_idx = _header_index(header_map, ["дата", "date", "день"], default_idx=0)
    point_idx = _header_index(header_map, ["пункт", "пвз", "точка", "адрес"], default_idx=1)
    amount_idx = _header_index(header_map, ["сумма", "стоимость", "итого", "amount"], default_idx=5)
    status_idx = _header_index(header_map, ["статус", "status"], default_idx=6)
    manager_idx = _header_index(header_map, ["фамилия менеджера", "менеджер", "фио", "фамилия", "сотрудник"], default_idx=7)

    if date_idx is None or amount_idx is None:
        raise RuntimeError("Не удалось найти обязательные колонки даты/суммы")

    out: list[DisputeXlsxRow] = []
    current_date: date | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[date_idx] if date_idx < len(row) else None
        parsed_date = parse_date(date_val)
        if parsed_date:
            current_date = parsed_date
        if current_date is None:
            continue

        point_raw = _cell_text(row[point_idx] if point_idx is not None and point_idx < len(row) else None)
        status_raw = _cell_text(row[status_idx] if status_idx is not None and status_idx < len(row) else None)
        manager_raw = _cell_text(row[manager_idx] if manager_idx is not None and manager_idx < len(row) else None)
        amount_raw = row[amount_idx] if amount_idx < len(row) else None
        amount_txt = _cell_text(amount_raw)

        # Skip service/template rows where only status is prefilled ("без статуса").
        if not amount_txt and not point_raw and not manager_raw:
            continue
        if not amount_txt:
            continue

        amount = parse_decimal(amount_txt, default=Decimal("0"))

        payload = {
            "date": current_date.isoformat(),
            "point": point_raw,
            "manager": manager_raw,
            "status": status_raw,
            "amount": str(amount),
        }

        out.append(
            DisputeXlsxRow(
                record_date=current_date,
                point_key=point_raw,
                manager_name=manager_raw,
                status=status_raw,
                amount_rub=amount,
                raw_payload=json.dumps(payload, ensure_ascii=False),
            )
        )

    return out


def _resolve_point_id(points: list[Point], raw_name: str | None) -> int | None:
    if not raw_name:
        return None
    needle = _extract_address_key(raw_name)
    if not needle:
        return None

    exact_map: dict[str, int] = {}
    for p in points:
        for source in (p.name, p.address, p.short_name):
            if source:
                exact_map[_extract_address_key(source)] = p.id

    exact = exact_map.get(needle)
    if exact:
        return exact

    for key, point_id in exact_map.items():
        if needle in key or key in needle:
            return point_id
    return None


def _resolve_user_id(user_map: dict[str, int], name: str | None) -> int | None:
    if not name:
        return None
    for key in _person_keys(name):
        if key in user_map:
            return user_map[key]
    return None


async def _build_user_name_map(session) -> dict[str, int]:
    result = await session.execute(select(User))
    users = result.scalars().all()
    out: dict[str, int] = {}
    for user in users:
        for key in _person_keys(user.full_name):
            out[key] = user.id
        if user.last_name:
            out[_norm_key(user.last_name)] = user.id
    return out


async def run_import(file_path: Path, sheet_name: str | None) -> None:
    rows = _parse_xlsx(file_path, sheet_name)
    if not rows:
        print("No valid dispute rows found in workbook.")
        return

    period_start = min(r.record_date for r in rows)
    period_end = max(r.record_date for r in rows)

    await init_db()
    async with SessionLocal() as session:
        user_map = await _build_user_name_map(session)
        points_result = await session.execute(select(Point))
        points = points_result.scalars().all()

        await session.execute(
            delete(MotivationRecord).where(
                MotivationRecord.source == MotivationSource.DISPUTE,
                MotivationRecord.record_date >= period_start,
                MotivationRecord.record_date <= period_end,
            )
        )

        records: list[MotivationRecord] = []
        unresolved_users: set[str] = set()
        unresolved_points: set[str] = set()

        for row in rows:
            user_id = _resolve_user_id(user_map, row.manager_name)
            point_id = _resolve_point_id(points, row.point_key)

            if row.manager_name and not user_id:
                unresolved_users.add(row.manager_name)
            if row.point_key and not point_id:
                unresolved_points.add(row.point_key)

            records.append(
                MotivationRecord(
                    source=MotivationSource.DISPUTE,
                    record_date=row.record_date,
                    point_id=point_id,
                    user_id=user_id,
                    manager_name=row.manager_name,
                    acceptance_amount_rub=Decimal("0"),
                    issued_items_count=0,
                    tickets_count=0,
                    disputed_amount_rub=row.amount_rub,
                    status=row.status,
                    raw_payload=row.raw_payload,
                )
            )

        session.add_all(records)
        await session.commit()

    print(f"Workbook: {file_path}")
    print(f"Inserted dispute rows: {len(rows)}")
    print(f"Period replaced: {period_start} .. {period_end}")
    print(f"Unresolved users: {len(unresolved_users)}")
    if unresolved_users:
        for name in sorted(unresolved_users):
            print(f"  - {name}")
    print(f"Unresolved points: {len(unresolved_points)}")
    if unresolved_points:
        for point in sorted(unresolved_points):
            print(f"  - {point}")


def _arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import disputes from XLSX into motivation_records")
    parser.add_argument("--file", required=True, help="Path to XLSX file")
    parser.add_argument("--sheet", default=None, help="Optional sheet name")
    return parser


async def _amain() -> None:
    args = _arg_parser().parse_args()
    await run_import(Path(args.file), args.sheet)


if __name__ == "__main__":
    asyncio.run(_amain())
