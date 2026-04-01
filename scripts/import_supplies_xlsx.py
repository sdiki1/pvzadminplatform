from __future__ import annotations

import argparse
import asyncio
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.db.models import Point, SupplyItem, SupplyRequestHeader, SupplyRequestItem, User
from app.db.session import SessionLocal, init_db
from app.utils.parsing import normalize_text, parse_date

IMPORT_TAG = "[excel_supplies]"

NICKNAMES = {
    "дима": "дмитрий",
    "даня": "даниил",
    "ксюша": "ксения",
    "даша": "дария",
    "дарья": "дария",
    "вика": "виктория",
    "варя": "варвара",
    "катя": "екатерина",
    "настя": "анастасия",
}


@dataclass
class ParsedLine:
    point_id: int
    request_date: date
    manager_name: str | None
    item_name: str
    raw_status_text: str
    item_status: str
    requested_qty: Decimal | None
    delivered_qty: Decimal | None


def _norm_key(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^\w\sа-яА-ЯёЁ]", " ", text, flags=re.UNICODE)
    text = normalize_text(text.replace("улица", "ул").replace("ул.", "ул"))
    return text


def _extract_address_key(value: str) -> str:
    text = _norm_key(value)
    parts = [
        p
        for p in text.split()
        if p
        not in {
            "пвз",
            "ул",
            "улица",
            "г",
            "город",
            "wb",
            "wildberries",
            "расходников",
            "заказ",
            "расходники",
        }
    ]
    return " ".join(parts)


def _person_tokens(name: str | None) -> list[str]:
    if not name:
        return []
    t = normalize_text(name).replace("ё", "е")
    t = re.sub(r"[^а-яa-z\s\-]", " ", t, flags=re.UNICODE)
    tokens = [x for x in t.split() if x]
    out: list[str] = []
    for tok in tokens:
        out.append(NICKNAMES.get(tok, tok))
    return out


def _person_keys(name: str | None) -> set[str]:
    tokens = _person_tokens(name)
    if not tokens:
        return set()

    keys = {" ".join(tokens), tokens[0], tokens[-1]}
    if len(tokens) >= 2:
        keys.add(f"{tokens[0]} {tokens[1]}")
        keys.add(f"{tokens[1]} {tokens[0]}")
        keys.add(f"{tokens[0]} {tokens[-1]}")
        keys.add(f"{tokens[-1]} {tokens[0]}")
    return keys


def _resolve_user_id(user_map: dict[str, int], raw_name: str | None) -> int | None:
    if not raw_name:
        return None
    for key in _person_keys(raw_name):
        if key in user_map:
            return user_map[key]
    return None


def _status_from_text(text: str) -> str:
    t = normalize_text(text)
    if not t:
        return "requested"
    if "отмен" in t:
        return "cancelled"
    if "все есть" in t or "всё есть" in t:
        return "in_stock"
    if "выдан" in t or "выда" in t or "привзли" in t:
        return "delivered"
    if "привез" in t or "привезем" in t or "привезём" in t or "перемещ" in t or "в пути" in t:
        return "in_transit"
    if "заказ" in t or "закза" in t:
        return "ordered"
    return "requested"


def _qty_from_text(text: str) -> Decimal | None:
    t = normalize_text(text).replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)\s*шт", t)
    if not m:
        return None
    try:
        return Decimal(m.group(1))
    except Exception:
        return None


def _header_status(line_statuses: list[str]) -> str:
    status_set = set(line_statuses)
    if not status_set:
        return "new"
    if status_set == {"in_stock"}:
        return "closed"
    if status_set.issubset({"cancelled"}):
        return "cancelled"
    if status_set.issubset({"delivered", "in_stock"}) and "delivered" in status_set:
        return "delivered"
    if "in_transit" in status_set:
        return "in_transit"
    if "ordered" in status_set or "requested" in status_set:
        if status_set == {"ordered"}:
            return "ordered"
        return "partially_ordered"
    return "new"


def _best_date_row(ws, max_row: int, max_col: int) -> int | None:
    best_row = None
    best_count = 0
    for row_idx in range(1, max_row + 1):
        cnt = 0
        for col_idx in range(1, max_col + 1):
            if parse_date(ws.cell(row_idx, col_idx).value):
                cnt += 1
        if cnt > best_count:
            best_count = cnt
            best_row = row_idx
    return best_row if best_count > 0 else None


def _manager_row(ws, date_row: int, max_row: int) -> int | None:
    for row_idx in range(date_row + 1, min(date_row + 10, max_row) + 1):
        v = ws.cell(row_idx, 1).value
        if v and "заполнял" in normalize_text(str(v)):
            return row_idx
    return None


def _extract_point_id(sheet_name: str, points_by_key: dict[str, int]) -> int | None:
    key = _extract_address_key(sheet_name)
    if key in points_by_key:
        return points_by_key[key]
    for pkey, point_id in points_by_key.items():
        if key in pkey or pkey in key:
            return point_id
    return None


def _parse_supply_sheets(
    workbook_path: Path,
    points_by_key: dict[str, int],
) -> tuple[list[ParsedLine], set[str], set[str]]:
    wb = load_workbook(workbook_path, data_only=True)
    parsed_lines: list[ParsedLine] = []
    unresolved_points: set[str] = set()
    all_item_names: set[str] = set()

    for sheet_name in wb.sheetnames:
        if "расход" not in normalize_text(sheet_name):
            continue

        ws = wb[sheet_name]
        max_row = min(ws.max_row, 1500)
        max_col = min(ws.max_column, 80)

        point_id = _extract_point_id(sheet_name, points_by_key)
        if not point_id:
            unresolved_points.add(sheet_name)
            continue

        date_row = _best_date_row(ws, max_row=max_row, max_col=max_col)
        if not date_row:
            continue

        mgr_row = _manager_row(ws, date_row, max_row=max_row)
        start_row = (mgr_row + 1) if mgr_row else (date_row + 1)

        date_by_col: dict[int, date] = {}
        for col_idx in range(2, max_col + 1):
            d = parse_date(ws.cell(date_row, col_idx).value)
            if d:
                date_by_col[col_idx] = d

        blank_streak = 0
        for row_idx in range(start_row, max_row + 1):
            raw_item = ws.cell(row_idx, 1).value
            item_name = str(raw_item).strip() if raw_item is not None else ""
            if not item_name:
                blank_streak += 1
                if blank_streak >= 120:
                    break
                continue
            blank_streak = 0

            if "заполнял" in normalize_text(item_name):
                continue

            all_item_names.add(item_name)

            for col_idx, req_date in date_by_col.items():
                raw_cell = ws.cell(row_idx, col_idx).value
                raw_text = str(raw_cell).strip() if raw_cell is not None else ""
                if not raw_text:
                    continue

                manager_name = None
                if mgr_row:
                    manager_cell = ws.cell(mgr_row, col_idx).value
                    manager_name = str(manager_cell).strip() if manager_cell is not None else None
                    if manager_name == "":
                        manager_name = None

                item_status = _status_from_text(raw_text)
                qty = _qty_from_text(raw_text)
                delivered_qty = qty if item_status == "delivered" else None

                parsed_lines.append(
                    ParsedLine(
                        point_id=point_id,
                        request_date=req_date,
                        manager_name=manager_name,
                        item_name=item_name,
                        raw_status_text=raw_text,
                        item_status=item_status,
                        requested_qty=qty,
                        delivered_qty=delivered_qty,
                    )
                )

    return parsed_lines, all_item_names, unresolved_points


async def _build_points_map(session) -> dict[str, int]:
    result = await session.execute(select(Point))
    points = result.scalars().all()

    out: dict[str, int] = {}
    for p in points:
        for source in [p.name, p.address, p.short_name]:
            if source:
                out[_extract_address_key(str(source))] = p.id
    return out


async def _build_user_map(session) -> dict[str, int]:
    result = await session.execute(select(User))
    users = result.scalars().all()
    out: dict[str, int] = {}
    for user in users:
        for key in _person_keys(user.full_name):
            out[key] = user.id
    return out


async def _get_wb_marketplace_id(session) -> int | None:
    from app.db.models import Marketplace

    result = await session.execute(select(Marketplace).where(Marketplace.code == "wb"))
    row = result.scalar_one_or_none()
    return row.id if row else None


async def run_import(file_path: Path) -> None:
    await init_db()
    async with SessionLocal() as session:
        points_by_key = await _build_points_map(session)
        user_map = await _build_user_map(session)
        wb_marketplace_id = await _get_wb_marketplace_id(session)

        parsed_lines, all_item_names, unresolved_points = _parse_supply_sheets(file_path, points_by_key)
        if not parsed_lines:
            print("No supply rows found to import.")
            if unresolved_points:
                print("Unresolved supply sheets:")
                for s in sorted(unresolved_points):
                    print(f"  - {s}")
            return

        # Ensure supply catalog items exist.
        existing_items_result = await session.execute(select(SupplyItem))
        existing_items = existing_items_result.scalars().all()
        item_by_norm: dict[str, SupplyItem] = {normalize_text(i.name): i for i in existing_items}
        created_items = 0
        for item_name in sorted(all_item_names):
            key = normalize_text(item_name)
            if key in item_by_norm:
                continue
            obj = SupplyItem(
                marketplace_id=wb_marketplace_id,
                category="Расходники",
                name=item_name,
                unit="шт",
                min_qty=None,
                is_active=True,
                comment=f"{IMPORT_TAG} from workbook",
            )
            session.add(obj)
            await session.flush()
            item_by_norm[key] = obj
            created_items += 1

        # Reimport guard: replace only previously imported workbook-based data.
        await session.execute(delete(SupplyRequestHeader).where(SupplyRequestHeader.comment.ilike(f"{IMPORT_TAG}%")))
        await session.commit()

        # Group rows by (point_id, request_date) -> header
        grouped: dict[tuple[int, date], list[ParsedLine]] = defaultdict(list)
        for line in parsed_lines:
            grouped[(line.point_id, line.request_date)].append(line)

        unresolved_managers: set[str] = set()
        created_headers = 0
        created_lines = 0

        for (point_id, req_date), lines in sorted(grouped.items(), key=lambda x: (x[0][1], x[0][0])):
            statuses = [ln.item_status for ln in lines]
            header = SupplyRequestHeader(
                point_id=point_id,
                request_date=req_date,
                requested_by_employee_id=None,  # set below
                status=_header_status(statuses),
                plan_delivery_date=None,
                actual_delivery_date=req_date if _header_status(statuses) in {"closed", "delivered"} else None,
                comment=f"{IMPORT_TAG} source={file_path.name}",
                created_by_user_id=None,
                updated_by_user_id=None,
            )

            # Most frequent manager name for this point/date.
            manager_candidates = [ln.manager_name for ln in lines if ln.manager_name]
            manager_name = manager_candidates[0] if manager_candidates else None
            if manager_name:
                user_id = _resolve_user_id(user_map, manager_name)
                if user_id:
                    header.requested_by_employee_id = user_id
                else:
                    unresolved_managers.add(manager_name)

            session.add(header)
            await session.flush()
            created_headers += 1

            # Upsert unique items within header by supply_item_id (in case duplicates in source).
            by_item: dict[int, ParsedLine] = {}
            for ln in lines:
                si = item_by_norm.get(normalize_text(ln.item_name))
                if not si:
                    continue
                by_item[si.id] = ln

            for supply_item_id, ln in by_item.items():
                line = SupplyRequestItem(
                    request_id=header.id,
                    supply_item_id=supply_item_id,
                    requested_qty=ln.requested_qty,
                    approved_qty=None,
                    delivered_qty=ln.delivered_qty,
                    item_status=ln.item_status,
                    status_date=req_date,
                    raw_status_text=ln.raw_status_text,
                    comment=None,
                )
                session.add(line)
                created_lines += 1

        await session.commit()

    min_date = min(ln.request_date for ln in parsed_lines)
    max_date = max(ln.request_date for ln in parsed_lines)
    print(f"Workbook: {file_path}")
    print(f"Imported headers: {created_headers}")
    print(f"Imported line items: {created_lines}")
    print(f"Created catalog items: {created_items}")
    print(f"Period: {min_date} .. {max_date}")
    print(f"Unresolved points: {len(unresolved_points)}")
    if unresolved_points:
        for s in sorted(unresolved_points):
            print(f"  - {s}")
    print(f"Unresolved managers: {len(unresolved_managers)}")
    if unresolved_managers:
        for n in sorted(unresolved_managers):
            print(f"  - {n}")


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import supplies requests from WB workbook")
    parser.add_argument("--file", required=True, help="Path to workbook .xlsx")
    return parser


async def _amain() -> None:
    args = _build_arg_parser().parse_args()
    await run_import(Path(args.file))


if __name__ == "__main__":
    asyncio.run(_amain())
